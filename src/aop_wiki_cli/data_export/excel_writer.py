"""
Generic Excel writing functions using field configurations.

This module provides reusable Excel writing functionality that works with
field configurations, similar to how csv_writer.py works for CSV exports.

Key features:
- Automatic sheet creation from field configurations
- Configurable field transformations (reuses CSV transformers)
- Built-in sorting support
- Column width optimization
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Union, Optional

from .field_definitions import FieldConfigSet
from .transformers import get_transformer
from .csv_writer import get_nested_value


def _transform_value(value: Any, transformer_name: Optional[str]) -> Any:
    """
    Transform a value using the specified transformer.
    
    Args:
        value: Raw value to transform
        transformer_name: Name of transformer function to use
        
    Returns:
        Transformed value
    """
    if transformer_name:
        transformer = get_transformer(transformer_name)
        if transformer:
            return transformer(value)
    return value


def write_excel_sheet(
    ws,
    data: Union[Dict, List[Dict]],
    field_config: FieldConfigSet,
    sort_key: Optional[str] = None,
    reverse_sort: bool = False,
    start_row: int = 1,
) -> int:
    """
    Write data to an Excel worksheet using field configuration.
    
    Args:
        ws: Worksheet object to write to
        data: Dictionary of items (keyed by ID) or list of items
        field_config: Field configuration set defining columns and transformations
        sort_key: Optional key to sort by (supports dot notation)
        reverse_sort: Whether to sort in descending order
        start_row: Row number to start writing (default 1 for headers)
        
    Returns:
        Last row number written
    """
    # Convert dict to list if needed
    if isinstance(data, dict):
        items = list(data.values())
    else:
        items = data
    
    if not items:
        return start_row
    
    # Sort if requested
    if sort_key:
        try:
            items = sorted(
                items,
                key=lambda x: get_nested_value(x, sort_key, 0),
                reverse=reverse_sort
            )
        except (TypeError, KeyError) as e:
            print(f"⚠ Warning: Could not sort by '{sort_key}': {e}")
    
    # Write headers
    current_row = start_row
    for col_idx, field_config_item in enumerate(field_config.fields, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=field_config_item.name)
        cell.font = Font(bold=True)
    
    current_row += 1
    
    # Write data rows
    for item in items:
        for col_idx, field_config_item in enumerate(field_config.fields, start=1):
            # Get value from item
            raw_value = get_nested_value(item, field_config_item.source_key, default='')
            
            # Transform value
            transformed_value = _transform_value(raw_value, field_config_item.transformer)
            
            # Write to cell
            ws.cell(row=current_row, column=col_idx, value=transformed_value)
        
        current_row += 1
    
    return current_row - 1  # Return last row written


def auto_size_columns(ws, field_config: FieldConfigSet, min_width: int = 10, max_width: int = 40):
    """
    Auto-size columns based on field configuration and content.
    
    Args:
        ws: Worksheet object
        field_config: Field configuration set
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for col_idx, field_config_item in enumerate(field_config.fields, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Base width on name length
        name_width = len(field_config_item.name) + 2
        width = max(min_width, min(name_width, max_width))
        
        # Adjust for specific transformers that might create longer content
        if field_config_item.transformer == 'join_list':
            width = max_width
        
        ws.column_dimensions[col_letter].width = width
        
        # Add text wrapping for wide columns
        if width >= 30:
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')


def write_excel_workbook(
    sheets: List[Dict[str, Any]],
    output_path: Union[str, Path],
    auto_filter: bool = True
) -> None:
    """
    Write multiple sheets to an Excel workbook.
    
    Args:
        sheets: List of sheet configurations, each containing:
            - name: Sheet name
            - data: Data to write
            - field_config: Field configuration for the sheet
            - sort_key: Optional sort key
            - reverse_sort: Optional reverse sort flag
        output_path: Path to output Excel file
        auto_filter: Whether to add auto-filter to sheets
    
    Example:
        sheets = [
            {
                'name': 'Events',
                'data': events_list,
                'field_config': EVENT_FIELDS_MIN,
                'sort_key': 'id'
            },
            {
                'name': 'Summary',
                'data': summary_list,
                'field_config': SUMMARY_FIELDS,
            }
        ]
        write_excel_workbook(sheets, outputs_dir() / 'report.xlsx')
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet_config in sheets:
        sheet_name = sheet_config['name'][:31]  # Excel limit
        ws = wb.create_sheet(sheet_name)
        
        # Write data to sheet
        last_row = write_excel_sheet(
            ws=ws,
            data=sheet_config['data'],
            field_config=sheet_config['field_config'],
            sort_key=sheet_config.get('sort_key'),
            reverse_sort=sheet_config.get('reverse_sort', False)
        )
        
        # Auto-size columns
        auto_size_columns(ws, sheet_config['field_config'])
        
        # Add auto-filter if requested
        if auto_filter and last_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(sheet_config['field_config'].fields))}{last_row}"
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel workbook written to {output_path} ({len(wb.sheetnames)} sheets, {sum(ws.max_row - 1 for ws in wb.worksheets)} total rows)")

