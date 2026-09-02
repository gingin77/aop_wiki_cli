"""
Module for exporting harmonized KER evidence tables to Excel workbooks.

This module provides functionality to transform harmonized KER evidence data
from JSON format into a multi-sheet Excel workbook with:
- Summary sheet with harmonization statistics
- Sheets for all harmonized and unmatched headers
- Sheet with all harmonizable KERs and their upstream/downstream KEs
- Individual sheets for each KER with evidence tables and metadata
"""

import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd
from datetime import date
import pprint as pp

from aop_wiki_cli.paths import outputs_dir
from aop_wiki_cli.utilities import get_single_aopwiki_url


def _create_sheet_for_aops(wb, summary, aops_selected_for_workbooks=None, aop_to_ker_dict=None, aop_info=None):
    """Create a worksheet listing all AOPs with harmonizable KERs."""
    harmonizable_kers = summary.get('harmonizable_kers', [])
    filtered_aop_to_ker_dict = {}

    for aop_id, ker_ids in aop_to_ker_dict.items():
        filtered_kers = [ker_id for ker_id in ker_ids if ker_id in harmonizable_kers]
        if filtered_kers:
            filtered_aop_to_ker_dict[aop_id] = filtered_kers

    aops_harm_evi = summary.get('aops_harm_evi', [])
    
    if not aops_harm_evi or aops_harm_evi == []:
        print("Warning: No AOPs with harmonizable KERs found in summary")
        return
    
    aop_sheet = wb.create_sheet("AOPs_with_harmonizable_KERs")
    headers = ["AOP ID", "AOP Title", "Selected for Workbook?", "OECD Status","Count Harmonizable KERs", "KER IDs"]

    for col_idx, header in enumerate(headers, start=1):
        aop_sheet.cell(row=1, column=col_idx, value=header)

    # Set column widths
    aop_sheet.column_dimensions['B'].width = 70  # AOP Title column    
    aop_sheet.column_dimensions['D'].width = 20  # OECD Status column
    
    # Build sortable dictionary combining AOP ID with KER count
    aop_ker_counts = {aop_id: len(filtered_aop_to_ker_dict.get(aop_id, [])) for aop_id in aops_harm_evi}
    sorted_aops = sorted(aop_ker_counts.items(), key=lambda x: x[1], reverse=True)
    
    for row_idx, (aop_id, ker_count) in enumerate(sorted_aops, start=2):
        aop_title = aop_info.get(aop_id, {}).get('title', 'N/A')
        oecd_status = aop_info.get(aop_id, {}).get('oecd_status', 'N/A')
        ker_ids = filtered_aop_to_ker_dict.get(aop_id, [])
        ker_ids_str = ", ".join(str(kid) for kid in ker_ids) if ker_ids else "N/A"
        
        # AOP ID with hyperlink
        aop_cell = aop_sheet.cell(row=row_idx, column=1, value=f"AOP {aop_id}")
        aop_url = get_single_aopwiki_url(aop_id, "aops")
        aop_cell.hyperlink = aop_url
        aop_cell.style = "Hyperlink"
        
        # AOP Title
        aop_sheet.cell(row=row_idx, column=2, value=aop_title)

        # Selected for Workbook?
        selected = "Y" if int(aop_id) in aops_selected_for_workbooks.keys() else ""
        aop_sheet.cell(row=row_idx, column=3, value=selected)
        
        # OECD Status
        aop_sheet.cell(row=row_idx, column=4, value=oecd_status)
        
        # Number of Harmonizable KERs
        aop_sheet.cell(row=row_idx, column=5, value=ker_count)
        
        # KER IDs
        aop_sheet.cell(row=row_idx, column=6, value=ker_ids_str)
    
    # Set text wrapping for title column (column B)
    for row in aop_sheet.iter_rows(min_col=2, max_col=2, min_row=1, max_row=aop_sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Apply auto-filter
    aop_sheet.auto_filter.ref = f"A1:F{aop_sheet.max_row}"

def _add_rows_from_field_type(all_rows, ker_id, field_data, field_type_name):
    """
    Helper function to add rows from a specific field type to the rows list.
    
    Args:
        all_rows: List to append rows to
        ker_id: KER identifier
        field_data: Dictionary containing 'tables' key with table data
        field_type_name: Name of field type (e.g., 'harmonized_weight_of_evidence')
    """
    tables = field_data.get('tables', {})
    for table_index, rows in tables.items():
        for row in rows:
            enhanced_row = {
                'KER ID': ker_id,
                'Field Type': field_type_name,
                'Table Index': table_index
            }
            enhanced_row.update(row)
            all_rows.append(enhanced_row)


def _write_ke_metadata(ws, current_row, ke_label, ke_data):
    """
    Helper function to write KE metadata to a worksheet row with optional hyperlink.
    
    Args:
        ws: Worksheet object
        current_row: Current row number
        ke_label: Label for the KE (e.g., "Upstream KE:")
        ke_data: Dictionary containing KE information (title, id)
    
    Returns:
        Updated current_row
    """
    ke_title = ke_data.get('title', 'N/A')
    ke_id = ke_data.get('id', 'N/A')
    
    ws.cell(row=current_row, column=1, value=ke_label)
    ws.cell(row=current_row, column=2, value=ke_title)
    
    # Add hyperlink to KE if ID is available
    if ke_id != 'N/A':
        ke_url = get_single_aopwiki_url(ke_id, "events")
        cell_with_link = ws.cell(row=current_row, column=2)
        cell_with_link.hyperlink = ke_url
        cell_with_link.style = "Hyperlink"
    
    return current_row + 1


def _create_summary_sheet(wb, summary):
    """Create and populate the summary sheet."""
    summary_sheet = wb.create_sheet("Summary")
    summary_df = pd.DataFrame.from_dict(summary['counts'], orient='index', columns=['Value'])
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            summary_sheet.cell(row=r_idx, column=c_idx, value=value)


def _create_header_list_sheets(wb, summary):
    """Create sheets for all harmonized and unmatched headers."""
    summary_lists = dict((k, v) for k, v in summary.items() if "_headers" in k)
    for list_name, items in summary_lists.items():
        list_sheet = wb.create_sheet(list_name)
        list_sheet.cell(row=1, column=1, value=list_name)
        for idx, item in enumerate(items, start=2):
            list_sheet.cell(row=idx, column=1, value=item)


def _create_harmonizable_kers_sheet(wb, harmonized_data, summary):
    """Create sheet with all harmonizable KERs and their upstream/downstream KEs."""
    harmonizable_kers = summary.get('harmonizable_kers', [])
    harmonizable_kers_sheet = wb.create_sheet("harmonizable_kers")
    
    harmonizable_ker_headers = ["KER ID", "Upstream KE", "Downstream KE"]
    for header_idx, header in enumerate(harmonizable_ker_headers, start=1):
        harmonizable_kers_sheet.cell(row=1, column=header_idx, value=header)

    for idx, ker_id in enumerate(harmonizable_kers, start=2):
        ker_cell = harmonizable_kers_sheet.cell(row=idx, column=1, value=f"KER {ker_id}")
        ker_url = get_single_aopwiki_url(ker_id, "relationships", page_anchor="link_support")
        ker_cell.hyperlink = ker_url
        ker_cell.style = "Hyperlink"
        
        ker_data = harmonized_data.get(ker_id, {})
        upstream_ke = ker_data.get('upstream_ke', {})
        upstream_title = upstream_ke.get('title', 'N/A')
        harmonizable_kers_sheet.cell(row=idx, column=2, value=upstream_title)
        
        downstream_ke = ker_data.get('downstream_ke', {})
        downstream_title = downstream_ke.get('title', 'N/A')
        harmonizable_kers_sheet.cell(row=idx, column=3, value=downstream_title)


def _create_ker_sheet(wb, ker_id, ker_data):
    """Create a worksheet for a single KER with evidence tables and metadata."""
    all_rows = []
    
    # Get harmonized tables from the new structure
    harmonized_tables = ker_data.get('harmonized_tables', {})
    
    # Process each field's harmonized tables
    for field_name, field_tables in harmonized_tables.items():
        _add_rows_from_field_type(all_rows, ker_id, {'tables': field_tables}, field_name)
    
    # Create worksheet if there are rows
    if not all_rows:
        return
    
    # Convert to DataFrame for easier Excel writing
    df = pd.DataFrame(all_rows)
    
    # Reorder columns to put KER ID, Field Type, and Table Index first
    cols = ['KER ID', 'Field Type', 'Table Index']
    cols.extend([c for c in df.columns if c not in cols])
    df = df[cols]
    
    # Create worksheet with KER ID as name (Excel sheet names have 31 char limit)
    sheet_name = f"KER_{ker_id}"[:31]
    ws = wb.create_sheet(sheet_name)
    
    # Write metadata at the top of the sheet
    current_row = 1
    
    # KER ID with AOP-Wiki link
    ker_url = get_single_aopwiki_url(ker_id, "relationships", page_anchor="link_support")
    cell_with_link = ws.cell(row=current_row, column=1, value=f"KER {ker_id}")
    cell_with_link.hyperlink = ker_url
    cell_with_link.style = "Hyperlink"
    current_row += 1
    
    # Write KE metadata
    upstream_ke = ker_data.get('upstream_ke', {})
    current_row = _write_ke_metadata(ws, current_row, "Upstream KE:", upstream_ke)
    
    downstream_ke = ker_data.get('downstream_ke', {})
    current_row = _write_ke_metadata(ws, current_row, "Downstream KE:", downstream_ke)
    
    # Add blank row before data
    current_row += 1
    
    # Write DataFrame to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Set column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15  # Default width   
        if col_name in ["Upstream Key Event", "Downstream Key Event"]:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 70     # Apply text wrapping to these columns
        
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=current_row, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

def transform_outputs_for_excel(harmonized_data, output_excel_path, aops_selected_for_workbooks=None, ker_pages_only=False, aop_to_ker_dict=None, aop_info=None):
    """
    Transform harmonized KER evidence tables from JSON into an Excel workbook.
    
    Creates one worksheet per KER ID containing harmonized evidence from both
    weight_of_evidence and empirical_support tables, with added columns for:
    - KER ID
    - field type weight_of_evidence or empirical_support
    - table_index
    
    Also includes KER metadata at the top of each sheet:
    - KER ID with link to AOP-Wiki
    - Upstream Key Event title, with link & KE ID
    - Downstream Key Event title, with link & KE ID
    
    Args:
        harmonized_data (dict): Harmonized KER evidence data
        output_excel_path (str): Path where the Excel workbook will be saved
        aops_selected_for_workbooks (dict): Optional dict of selected AOP IDs
        ker_pages_only (bool): If True, only create KER sheets, skip summary
        aop_to_ker_dict (dict): Optional pre-collected AOP-to-KER mapping
        aop_info (dict): Optional pre-collected AOP information
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    summary = harmonized_data.get('summary', {})
    
    # Create summary and list sheets
    if not ker_pages_only:
        _create_summary_sheet(wb, summary)
        _create_sheet_for_aops(wb, summary, aops_selected_for_workbooks, aop_to_ker_dict, aop_info)
        _create_header_list_sheets(wb, summary)
        _create_harmonizable_kers_sheet(wb, harmonized_data, summary)

    # Create individual sheets for each KER
    ker_dataset = dict(sorted(((k, v) for k, v in harmonized_data.items() if k != 'summary'), key=lambda x: int(x[0])))
    for ker_id, ker_data in ker_dataset.items():
        _create_ker_sheet(wb, ker_id, ker_data)

    # Save workbook
    wb.save(output_excel_path)
    print(f"Excel workbook saved to {output_excel_path} with {len(wb.sheetnames)} worksheets")
    
if __name__ == "__main__":
    # Example usage
    tabulated_ker_dir = outputs_dir("tabulated_ker_evidence")
    harmonized_ker_json_path = tabulated_ker_dir / "harmonized_ker_evidence_2025-12-22.json"
    harmonized_ker_data = json.load(open(harmonized_ker_json_path, 'r', encoding='utf-8'))
    
    output_excel_path = tabulated_ker_dir / "harmonized_ker_evidence_2025-12-22.xlsx"
    transform_outputs_for_excel(harmonized_ker_data, output_excel_path)